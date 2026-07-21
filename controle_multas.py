"""Regras de negócio da aba CONTROLE DE MULTAS + gravação das multas novas.

Mapeia uma infração vinda do DETRAN para uma linha da aba CONTROLE DE MULTAS
(colunas B–F), decide o que é "pendente" e deduplica pelo Número do AI (coluna D).

Colunas da aba CONTROLE DE MULTAS:
    A=TRELLO  B=CARRO  C=DATA DA INFRAÇÃO  D=NÚMERO DO AI  E=ORGÃO AUTUADOR
    F=VALOR   G+=status/motorista/...

Sinks de escrita:
    - XlsxSink: grava num arquivo .xlsx (usado agora, para testes).
    - (futuro) SheetsSink: append direto no Google Sheets via Service Account.
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime

import openpyxl

SHEET_CONTROLE = "CONTROLE DE MULTAS"
SHEET_ORGAO = "ORGÃO AUTUADOR"

COL_B_CARRO = 2
COL_C_DATA = 3
COL_D_AIT = 4
COL_E_ORGAO = 5
COL_F_VALOR = 6


@dataclass
class MultaRow:
    carro: str          # B
    data: date | str    # C
    ait: str            # D
    orgao: str          # E
    valor: float | None  # F

    def to_cells(self) -> list:
        """Linha A..F para append no Google Sheets (A vazia; datas em dd/mm/aaaa)."""
        data = self.data.strftime("%d/%m/%Y") if isinstance(self.data, date) else (self.data or "")
        return ["", self.carro, data, self.ait, self.orgao,
                self.valor if self.valor is not None else ""]


# ---------------------------------------------------------------- parsing/regras

def is_pendente(inf: dict) -> bool:
    """True para multas 'não pagas / a vencer' (exclui Paga, anuladas e sem grupo)."""
    g = (inf.get("grupo") or "").strip().lower()
    if not g:
        return False
    if "não paga" in g or "nao paga" in g:
        return True
    if "vencer" in g or "vencid" in g:  # A Vencer / Vencida(s)
        return True
    return False


def parse_valor(valor) -> float | None:
    """'R$ 130,16' -> 130.16 ; '1.234,56' -> 1234.56."""
    if valor is None:
        return None
    s = str(valor).replace("R$", "").strip().replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_data(valor) -> date | str:
    """'12/02/2026' -> date(2026, 2, 12) ; devolve a string original se não parsear."""
    if valor is None:
        return ""
    try:
        return datetime.strptime(str(valor).strip(), "%d/%m/%Y").date()
    except ValueError:
        return str(valor).strip()


def orgao_str(cod_oat, orgao_fiscalizador, lookup: dict[str, str]) -> str:
    """Formata o órgão como na planilha ('000100 - PRF'), usando a aba ORGÃO AUTUADOR."""
    if cod_oat is None:
        return orgao_fiscalizador or ""
    code6 = f"{int(cod_oat):06d}"
    if code6 in lookup:
        return lookup[code6]
    return f"{code6} - {orgao_fiscalizador}".strip() if orgao_fiscalizador else code6


def build_row(carro: str, inf: dict, lookup: dict[str, str]) -> MultaRow:
    return MultaRow(
        carro=carro,
        data=parse_data(inf.get("data")),
        ait=str(inf.get("serieAIT") or "").strip(),
        orgao=orgao_str(inf.get("codOat"), inf.get("orgaoFiscalizador"), lookup),
        valor=parse_valor(inf.get("valor")),
    )


# ---------------------------------------------------------------- leitura da base

def existing_ait_from_rows(rows) -> set[str]:
    """Conjunto dos Números do AI (coluna D) a partir de linhas (xlsx ou Sheets)."""
    existing: set[str] = set()
    for i, row in enumerate(rows):
        if i == 0:
            continue
        ait = row[COL_D_AIT - 1] if len(row) >= COL_D_AIT else None
        if ait:
            existing.add(str(ait).strip().upper())
    return existing


def orgao_lookup_from_rows(rows) -> dict[str, str]:
    """Mapa '000100' -> '000100 - PRF' a partir das linhas da aba ORGÃO AUTUADOR."""
    lookup: dict[str, str] = {}
    for row in rows:
        a = row[0] if row and len(row) else None
        if not a:
            continue
        s = str(a).strip()
        m = re.match(r"^(\d{6})", s)
        if m:
            lookup[m.group(1)] = s
    return lookup


def read_existing_ait(path) -> set[str]:
    """Conjunto dos Números do AI (coluna D) já presentes na aba CONTROLE DE MULTAS (xlsx)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    existing = existing_ait_from_rows(wb[SHEET_CONTROLE].iter_rows(values_only=True))
    wb.close()
    return existing


def read_orgao_lookup(path) -> dict[str, str]:
    """Mapa '000100' -> '000100 - PRF' a partir da aba ORGÃO AUTUADOR (xlsx)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    lookup = orgao_lookup_from_rows(wb[SHEET_ORGAO].iter_rows(values_only=True))
    wb.close()
    return lookup


# ---------------------------------------------------------------- sink (escrita)

class XlsxSink:
    """Acumula linhas e grava num arquivo .xlsx (cópia da planilha base)."""

    def __init__(self, src_path, out_path):
        self.out_path = out_path
        self.wb = openpyxl.load_workbook(src_path)
        self.ws = self.wb[SHEET_CONTROLE]
        self._row = self._last_data_row()
        self.count = 0

    def _last_data_row(self) -> int:
        last = 1
        for r in range(1, self.ws.max_row + 1):
            if self.ws.cell(r, COL_B_CARRO).value or self.ws.cell(r, COL_D_AIT).value:
                last = r
        return last

    def append(self, row: MultaRow) -> None:
        self._row += 1
        r = self._row
        self.ws.cell(r, COL_B_CARRO, row.carro)
        c = self.ws.cell(r, COL_C_DATA, row.data)
        if isinstance(row.data, date):
            # minúsculas: o Google Sheets só reconhece d/m/y assim; com "DD/MM/YYYY"
            # ele renderiza literalmente "DD/06/YYYY" ao abrir o xlsx.
            c.number_format = "dd/mm/yyyy"
        self.ws.cell(r, COL_D_AIT, row.ait)
        self.ws.cell(r, COL_E_ORGAO, row.orgao)
        f = self.ws.cell(r, COL_F_VALOR, row.valor)
        f.number_format = "0.00"
        self.count += 1

    def save(self) -> None:
        self.wb.save(self.out_path)
