"""Fonte de dados dos veículos.

Hoje lê de um arquivo .xlsx local. A assinatura de `read_vehicles()` é a interface
estável — no futuro pode ser reimplementada para ler do Google Sheets (Service Account)
sem que o restante do programa mude.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl


SHEET_NAME = "LISTA DE CARROS"
# Colunas (1-indexadas) na aba LISTA DE CARROS.
COL_CARRO = 1    # A
COL_PLACA = 2    # B
COL_RENAVAM = 3  # C


@dataclass
class Vehicle:
    carro: str
    placa: str
    renavam: str


def _normalize_placa(value) -> str:
    if value is None:
        return ""
    return str(value).strip().upper().replace("-", "").replace(" ", "")


def _normalize_renavam(value) -> str:
    """RENAVAM vem do xlsx como float em notação científica (ex.: 1.134971084E9).

    int(float(...)) recupera o inteiro (1134971084), que é a mesma forma usada nas
    URLs públicas da planilha.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
    try:
        return str(int(float(value)))
    except (TypeError, ValueError):
        return str(value).strip()


def vehicles_from_rows(rows, limit: int | None = None) -> list[Vehicle]:
    """Constrói a lista de veículos a partir de linhas (xlsx ou Google Sheets).

    `rows` é um iterável de linhas, cada uma indexável por coluna (0-based A,B,C…).
    A primeira linha é o cabeçalho.
    """
    vehicles: list[Vehicle] = []
    for i, row in enumerate(rows):
        if i == 0:
            continue  # cabeçalho
        placa = _normalize_placa(row[COL_PLACA - 1] if len(row) >= COL_PLACA else None)
        renavam = _normalize_renavam(row[COL_RENAVAM - 1] if len(row) >= COL_RENAVAM else None)
        carro = str(row[COL_CARRO - 1] or "").strip() if len(row) >= COL_CARRO else ""
        if not placa or not renavam:
            continue
        vehicles.append(Vehicle(carro=carro, placa=placa, renavam=renavam))
        if limit is not None and len(vehicles) >= limit:
            break
    return vehicles


def read_vehicles(path: str | Path, limit: int | None = None) -> list[Vehicle]:
    """Lê a aba LISTA DE CARROS de um .xlsx e devolve os veículos válidos."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
    vehicles = vehicles_from_rows(ws.iter_rows(values_only=True), limit=limit)
    wb.close()
    return vehicles
